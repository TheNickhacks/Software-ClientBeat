from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from apps.geo.models import Region, Provincia, Comuna


class Command(BaseCommand):
    help = (
        'Carga / actualiza Regiones (16), Provincias y Comunas (346) desde el archivo '
        'regiones-chile.xlsx ubicado en el ROOT del proyecto Client Beat. '
        'Usa --reset para borrar toda la data geo existente ANTES de cargar.'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--reset',
            action='store_true',
            help='Borra TODAS las Comunas > Provincias > Regiones antes de insertar.',
        )

    def handle(self, *args, **options):
        reset = options.get('reset')

        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl no esta instalado. Ejecuta: pip install openpyxl')

        PROJECT_ROOT = settings.BASE_DIR.parent  # backend/.. = ClienBeat (raiz repo)

        import os
        found_xlsx = [f for f in os.listdir(PROJECT_ROOT) if f.startswith('regiones') and f.endswith('.xlsx')]
        if not found_xlsx:
            raise CommandError(f'No se encontro regiones-chile.xlsx en ROOT={PROJECT_ROOT}')

        path_xlsx = os.path.join(PROJECT_ROOT, found_xlsx[0])
        self.stdout.write(f'[INFO] Leyendo excel: {found_xlsx[0]}')

        if reset:
            self.stdout.write('[INFO] Reset solicitado: borrando Comunas -> Provincias -> Regiones...')
            Comuna.objects.all().delete()
            Provincia.objects.all().delete()
            Region.objects.all().delete()

        wb = load_workbook(path_xlsx, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        self.stdout.write(f'[INFO] Filas totales (incl headers malos): {len(all_rows)}')

        data_rows = []
        header_real = None
        for i, r in enumerate(all_rows):
            if i == 0:
                continue
            if i == 1:
                header_real = r
                self.stdout.write(f'[INFO] Header detectado fila {i}: {header_real}')
                continue
            if not r or not r[0]:
                continue
            if str(r[0]).strip() == str(header_real[0]).strip():
                continue
            data_rows.append(r)

        self.stdout.write(f'[INFO] Filas de datos utiles: {len(data_rows)}')

        region_cache = {}
        provincia_cache = {}

        total_regiones_nuevas = 0
        total_provincias_nuevas = 0
        total_comunas_nuevas = 0
        total_updates = 0

        orden_region_global = 0
        for row in data_rows:
            region_nombre = str(row[0]).strip() if row[0] else ''
            codigo_romano = str(row[1]).strip() if row[1] else ''
            provincia_nombre = str(row[2]).strip() if row[2] else ''
            comuna_nombre = str(row[3]).strip() if row[3] else ''
            if not region_nombre or not provincia_nombre or not comuna_nombre:
                continue

            if region_nombre not in region_cache:
                orden_region_global += 1
                reg_orden = orden_region_global
                region, created = Region.objects.get_or_create(
                    nombre=region_nombre,
                    defaults={
                        'codigo_romano': codigo_romano or f'R{orden_region_global}',
                        'orden': reg_orden,
                        'activo': True,
                    }
                )
                if not created:
                    need_upd = False
                    if codigo_romano and region.codigo_romano != codigo_romano:
                        region.codigo_romano = codigo_romano
                        need_upd = True
                    if region.orden != reg_orden:
                        region.orden = reg_orden
                        need_upd = True
                    if not region.activo:
                        region.activo = True
                        need_upd = True
                    if need_upd:
                        region.save()
                        total_updates += 1
                else:
                    total_regiones_nuevas += 1
                region_cache[region_nombre] = region
            region = region_cache[region_nombre]

            prov_key = (region_nombre, provincia_nombre)
            if prov_key not in provincia_cache:
                provincia, created = Provincia.objects.get_or_create(
                    region=region,
                    nombre=provincia_nombre,
                    defaults={
                        'orden': region.provincias.count() + 1,
                    }
                )
                if created:
                    total_provincias_nuevas += 1
                provincia_cache[prov_key] = provincia
            provincia = provincia_cache[prov_key]

            comuna, created = Comuna.objects.get_or_create(
                provincia=provincia,
                nombre=comuna_nombre,
                defaults={
                    'orden': provincia.comunas.count() + 1,
                }
            )
            if created:
                total_comunas_nuevas += 1

        wb.close()

        self.stdout.write('=' * 70)
        self.stdout.write('   RESULTADO CARGA GEOGRAFIA CHILE')
        self.stdout.write('=' * 70)
        self.stdout.write(f'   Regiones creadas:     {total_regiones_nuevas}')
        self.stdout.write(f'   Provincias creadas:   {total_provincias_nuevas}')
        self.stdout.write(f'   Comunas creadas:      {total_comunas_nuevas}')
        self.stdout.write(f'   Actualizadas:         {total_updates}')
        self.stdout.write(f'   --- Totales en BD ---')
        self.stdout.write(f'   Regiones totales:     {Region.objects.count()}')
        self.stdout.write(f'   Provincias totales:   {Provincia.objects.count()}')
        self.stdout.write(f'   Comunas totales:      {Comuna.objects.count()}')
        self.stdout.write('=' * 70)
