from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.conf import settings
from datetime import datetime

from apps.accounts.onboarding_service import onboarding_pendiente, onboarding_siguiente_paso


TERMINOS_Y_CONDICIONES = """
1. OBJETO Y ACEPTACIÓN
Client Beat es una plataforma de software como servicio (SaaS) que permite a dueños de negocios monitorear, en un panel centralizado, su reputación digital, compararla con la de sus competidores directos y capturar la opinión de sus propios clientes en el punto de venta, sin requerir conocimientos técnicos.
Al registrarse, acceder o utilizar el Servicio, el Cliente acepta íntegra y vinculantemente estos Términos y Condiciones. Si el Cliente no está de acuerdo con ellos, deberá abstenerse de registrarse o de utilizar la plataforma.

2. DEFINICIONES
Cliente: persona natural o jurídica que se registra y contrata el Servicio para su negocio.
Servicio o Plataforma: Client Beat, incluyendo su sitio web, panel de administración, motor de reputación y benchmarking, módulo de encuestas por código QR, motor de análisis de sentimiento y pasarela de pagos integrada.
Tenant: la instancia lógica y aislada de datos correspondiente a cada Cliente dentro de la arquitectura multi-tenant de la Plataforma.

3. DESCRIPCIÓN DEL SERVICIO (PLAN MVP)
El plan inicial contratado comprende:
• Reputación y benchmarking de competencia mediante Google Places API.
• Captura de voz directa mediante encuestas QR (NPS y CSAT).
• Motor de análisis de sentimiento léxico con modismos chilenos.
• Ciclo completo de suscripción y pago a través de Flow.

4. REGISTRO Y CUENTA
El Cliente debe ser mayor de 18 años y declarar esta condición al registrarse.
Client Beat no almacena contraseñas en texto plano ni tiene acceso a la contraseña del Cliente.
El Cliente es responsable de la confidencialidad de sus credenciales.

5. CONDICIONES ECONÓMICAS Y FACTURACIÓN
Los precios se expresan en UF o CLP, netos de IVA, conforme a la práctica comercial chilena.
La suscripción se renueva automáticamente al vencimiento de cada ciclo de facturación salvo cancelación oportuna.
En caso de mora, el acceso a la Plataforma podrá suspenderse.

6. PASARELA DE PAGOS FLOW
El cobro se procesa exclusivamente en los servidores de Flow. Client Beat no almacena datos de tarjetas.
Las notificaciones de pago se validan mediante firma digital antes de activar suscripciones.

7. PROPIEDAD INTELECTUAL
Client Beat conserva todos los derechos sobre su software, algoritmos y diseño. El Cliente conserva la titularidad de sus propios datos.

8. DISPONIBILIDAD Y RESPONSABILIDAD
El Servicio se presta "tal como está". La responsabilidad de Client Beat se limita al monto pagado en los últimos 12 meses.

9. PROTECCIÓN DE DATOS PERSONALES
Client Beat SpA (RUT 77.607.311-3) actúa como responsable del tratamiento. Los titulares pueden ejercer sus derechos ARCOPB al correo contacto@avaapp.cl.

10. LEY APLICABLE Y JURISDICCIÓN
Estos Términos se rigen por las leyes de la República de Chile y se someten a la jurisdicción de los tribunales ordinarios justicia.

Versión inicial — Agosto 2026.
"""

POLITICA_PRIVACIDAD = """
1. RESPONSABLE DEL TRATAMIENTO
Nombre: Client Beat SpA
RUT: 77.607.311-3
Domicilio: Asturias 790 E, Rancagua
Correo contacto: contacto@avaapp.cl

2. CATEGORÍAS DE DATOS
• Datos del Cliente: nombre o razón social, RUT, rubro, datos de contacto, credenciales, facturación.
• Datos de clientes finales: respuestas de satisfacción, comentarios, calificaciones, reseñas públicas de Google.
• Datos de navegación: registros técnicos de acceso, direcciones IP, logs de actividad.

3. FINALIDAD DEL TRATAMIENTO
• Gestión del registro y la relación contractual.
• Prestación del Servicio: reputación, benchmarking, encuestas, análisis de sentimiento.
• Cumplimiento de obligaciones legales y tributarias.

4. BASE DE LICITUD
El tratamiento se funda en la ejecución del contrato (servicio contratado), el consentimiento del titular (encuestas QR) y el interés legítimo del responsable (benchmarking anonimizado).

5. DERECHOS DEL TITULAR (ARCOPB)
Todo titular puede ejercer gratuitamente sus derechos de Acceso, Rectificación, Cancelación/Supresión, Oposición, Portabilidad y Bloqueo escribiendo a contacto@avaapp.cl. Plazo de respuesta: 30 días, prorrogables por 30 más.

6. CONSERVACIÓN
• Datos tributarios: 6 años.
• Logs de navegación: 12 meses.
• Datos de encuestas: hasta que el titular solicite su supresión.

7. TRANSFERENCIAS INTERNACIONALES
Neon (PostgreSQL), Google Places API y proveedores de correo pueden implicar transferencias internacionales. Se adoptan las garantías legales correspondientes.

8. SEGURIDAD
Aislamiento lógico multi-tenant, cifrado SSL/TLS, hashing de credenciales, respaldos periódicos y procedimiento de notificación de brechas.

9. NOTIFICACIÓN DE BRECHAS
En caso de vulnerabilidad relevante, se notificará a la Agencia de Protección de Datos y a los titulares afectados sin dilación indebida.
"""

POLITICA_COOKIES = """
¿Qué son las cookies?
Las cookies son pequeños archivos de texto que se almacenan en tu navegador al visitar una página web. Nos permiten recordar tus preferencias y mejorar la experiencia.

Cookies que utilizamos en Client Beat:
1. Cookies esenciales: necesarias para el funcionamiento del sitio (sesión de usuario, carrito, seguridad CSRF). No pueden desactivarse.
2. Cookies analíticas (opcionales): para medir el tráfico y mejorar el servicio.

Cómo gestionar las cookies
Puedes eliminar o bloquear las cookies desde la configuración de tu navegador. Ten en cuenta que algunas funcionalidades del sitio podrían dejar de funcionar correctamente.

Al continuar navegando por www.clientbeat.cl, aceptas nuestra política de cookies.
"""


_CONTENIDOS_LEGALES = {
    'terminos': {
        'titulo': 'Términos y Condiciones de Servicio',
        'fecha': '04 de Agosto del 2026',
        'contenido': TERMINOS_Y_CONDICIONES.strip(),
    },
    'privacidad': {
        'titulo': 'Política de Privacidad y Protección de Datos',
        'fecha': '04 de Agosto del 2026',
        'contenido': POLITICA_PRIVACIDAD.strip(),
    },
    'cookies': {
        'titulo': 'Política de Cookies',
        'fecha': '04 de Agosto del 2026',
        'contenido': POLITICA_COOKIES.strip(),
    },
}


TEMPORALIDAD_CHOICES = [
    ('SEMANAL', 'Semanal (7 días)', 7),
    ('MENSUAL', 'Mensual (30 días)', 30),
    ('TRIMESTRAL', 'Trimestral (90 días)', 90),
    ('SEMESTRAL', 'Semestral (180 días)', 180),
    ('ANUAL', 'Anual (365 días)', 365),
]


UBICACION_CHOICES = [
    ('KM_5',     'Hasta 5 km',     5),
    ('KM_10',    'Hasta 10 km',    10),
    ('KM_MAS_10','Más de 10 km',   9999),
    ('COMUNAL',  'Comunal',        None),
    ('REGIONAL', 'Regional',       None),
    ('NACIONAL', 'Nacional',       None),
]


NPS_CAT_CHOICES = [
    ('TODOS', 'Todas las categorías'),
    ('PROMOTOR', 'Solo Promotores (NPS 9-10)'),
    ('PASIVO', 'Solo Pasivos (NPS 7-8)'),
    ('DETRACTOR', 'Solo Detractores (NPS 0-6)'),
]


DIMENSIONES_TEMATICAS = {
    'ATENCION': {
        'nombre': 'Atención al Cliente',
        'icono': 'fa-user-tie',
        'color': 'cbblue',
        'keywords': ['atención', 'atencion', 'personal', 'trato', 'amable', 'amabilidad',
                     'rápido', 'rapido', 'velocidad', 'espera', 'demora', 'lento',
                     'caja', 'pago', 'horario', 'empleado', 'mozo', 'asesor', 'asesor',
                     'ayuda', 'amigable', 'cordial', 'grosero', 'mala atención', 'mal atendido'],
    },
    'PRODUCTO': {
        'nombre': 'Producto / Servicio',
        'icono': 'fa-mug-hot',
        'color': 'amber',
        'keywords': ['producto', 'comida', 'bebida', 'café', 'cafe', 'alimento', 'calidad',
                     'precio', 'variedad', 'fresco', 'frescura', 'presentación', 'presentacion',
                     'rico', 'sabor', 'soso', 'malo', 'barato', 'caro', 'cantidad', 'porción',
                     'porcion', 'servicio', 'resultado', 'stock', 'disponible', 'talla', 'modelo'],
    },
    'ESPACIO': {
        'nombre': 'Espacio / Ambiente',
        'icono': 'fa-couch',
        'color': 'purple',
        'keywords': ['local', 'ambiente', 'lugar', 'espacio', 'música', 'musica', 'decoración',
                     'decoracion', 'iluminación', 'iluminacion', 'temperatura', 'acogedor',
                     'cómodo', 'comodo', 'mobiliario', 'mesa', 'silla', 'sucio', 'desorden',
                     'orden', 'grande', 'pequeño', 'pequeno', 'ruido', 'silencio', 'estacionamiento'],
    },
    'LIMPIEZA': {
        'nombre': 'Limpieza e Higiene',
        'icono': 'fa-soap',
        'color': 'emerald',
        'keywords': ['limpio', 'limpieza', 'aseo', 'higiene', 'baño', 'bano', 'aseado',
                     'sucio', 'basura', 'mancha', 'olor', 'desinfectado', 'utensilios',
                     'vajilla', 'mesa', 'piso', 'orden', 'cocina', 'higienico', 'sanitario'],
    },
}


def landing(request):
    return render(request, 'landing.html')


def _parse_date(s):
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _analizar_tematicas(comentarios):
    """Recibe lista de textos y retorna conteo de menciones por dimensión temática."""
    result = {k: {'menciones': 0, 'nombre': v['nombre'], 'icono': v['icono'], 'color': v['color']}
              for k, v in DIMENSIONES_TEMATICAS.items()}
    total = 0
    for texto in comentarios:
        if not texto:
            continue
        txt = texto.lower()
        matched_any = False
        for key, dim in DIMENSIONES_TEMATICAS.items():
            if any(k in txt for k in dim['keywords']):
                result[key]['menciones'] += 1
                total += 1
                matched_any = True
        if not matched_any:
            pass
    for key in result:
        if total > 0:
            result[key]['pct'] = round(100 * result[key]['menciones'] / total, 1)
        else:
            result[key]['pct'] = 0
    return result, total


def _estrellas_html(n):
    n = max(0, min(5, int(n or 0)))
    return ('★' * n) + ('☆' * (5 - n))


@login_required(login_url='/accounts/login/')
def dashboard(request):
    if onboarding_pendiente(request.user):
        paso, _ = onboarding_siguiente_paso(request.user)
        return redirect(f'/accounts/onboarding/?paso={paso}')

    # ========= TABS / PESTAÑAS =========
    tabs = [
        {'key': 'resumen',  'label': '📊 Resumen KPI',      'icon': 'fa-chart-column'},
        {'key': 'encuestas','label': '✅ Encuestas QR',      'icon': 'fa-square-check'},
        {'key': 'google',   'label': '⭐ Reseñas Google',    'icon': 'fa-star'},
        {'key': 'analisis', 'label': '🧠 Análisis Latencia & Temática', 'icon': 'fa-brain'},
        {'key': 'benchmark','label': '🏆 Benchmark',         'icon': 'fa-trophy'},
    ]
    tab_actual = request.GET.get('tab', 'resumen')
    if tab_actual not in [t['key'] for t in tabs]:
        tab_actual = 'resumen'

    # ========= FILTROS GLOBALES =========
    temp_key = request.GET.get('temporalidad', 'MENSUAL')
    temp_map = {c[0]: c for c in TEMPORALIDAD_CHOICES}
    temp_selected = temp_map.get(temp_key, TEMPORALIDAD_CHOICES[1])
    temp_days = temp_selected[2]
    temp_label = temp_selected[1]

    ubi_key = request.GET.get('ubicacion', 'COMUNAL')
    ubi_map = {c[0]: c for c in UBICACION_CHOICES}
    ubi_selected = ubi_map.get(ubi_key, UBICACION_CHOICES[3])
    ubi_km = ubi_selected[2]
    ubi_label = ubi_selected[1]

    filtro_local_id = request.GET.get('local')
    filtro_nps_cat = request.GET.get('nps_cat', 'TODOS')
    filtro_fecha_desde = _parse_date(request.GET.get('fecha_desde'))
    filtro_fecha_hasta = _parse_date(request.GET.get('fecha_hasta'))
    filtro_fecha_exacta = _parse_date(request.GET.get('fecha_exacta'))

    ctx = {
        'welcome': request.GET.get('welcome') == '1',
        'negocio': getattr(request, 'negocio', None),
        'negocios_qs': getattr(request, 'negocios_qs', None),
        'rol_actual': getattr(request, 'rol_actual_negocio', None),
        'temporalidad_actual': temp_key,
        'temporalidad_label': temp_label,
        'temporalidad_dias': temp_days,
        'temporalidad_choices': TEMPORALIDAD_CHOICES,
        'ubicacion_actual': ubi_key,
        'ubicacion_label': ubi_label,
        'ubicacion_km': ubi_km,
        'ubicacion_choices': UBICACION_CHOICES,
        'tabs': tabs,
        'tab_actual': tab_actual,
        'filtro_nps_cat': filtro_nps_cat,
        'nps_cat_choices': NPS_CAT_CHOICES,
        'filtro_local_id': filtro_local_id,
        'filtro_fecha_desde': request.GET.get('fecha_desde', '') if filtro_fecha_desde is None else filtro_fecha_desde.strftime('%Y-%m-%d'),
        'filtro_fecha_hasta': request.GET.get('fecha_hasta', '') if filtro_fecha_hasta is None else filtro_fecha_hasta.strftime('%Y-%m-%d'),
        'filtro_fecha_exacta': request.GET.get('fecha_exacta', '') if filtro_fecha_exacta is None else filtro_fecha_exacta.strftime('%Y-%m-%d'),
        'filtro_fecha_desde_obj': filtro_fecha_desde,
        'filtro_fecha_hasta_obj': filtro_fecha_hasta,
        'filtro_fecha_exacta_obj': filtro_fecha_exacta,
    }
    negocio = ctx['negocio']
    if negocio is not None:
        from apps.billing.models import Suscripcion, EstadoSuscripcionChoices
        suscripcion = (
            Suscripcion.objects
            .filter(negocio=negocio, estado=EstadoSuscripcionChoices.ACTIVA)
            .select_related('plan')
            .order_by('-fecha_inicio')
            .first()
        )
        ctx['suscripcion_activa'] = suscripcion
        ctx['plan_activo'] = suscripcion.plan if suscripcion else None

        ctx['locales_disponibles'] = list(
            negocio.locales.filter(estado='ACTIVO').values('id', 'nombre').order_by('nombre')
        )

        from django.db.models import Count, Avg, Q, IntegerField, Sum, Case, When, F
        from apps.encuestas.models import RespuestaEncuesta, EmocionCSATChoices
        from django.utils import timezone
        desde_temp = timezone.now() - timezone.timedelta(days=temp_days)

        # Construir base QS con filtros
        locales_negocio_qs = RespuestaEncuesta.objects.filter(
            local__negocio=negocio,
        ).select_related('local', 'plantilla')

        if filtro_local_id:
            try:
                lid = int(filtro_local_id)
                locales_negocio_qs = locales_negocio_qs.filter(local_id=lid)
            except (ValueError, TypeError):
                pass
        if filtro_fecha_desde:
            dt_desde = datetime.combine(filtro_fecha_desde, datetime.min.time(), tzinfo=timezone.get_current_timezone())
            locales_negocio_qs = locales_negocio_qs.filter(fecha_respuesta__gte=dt_desde)
        if filtro_fecha_hasta:
            dt_hasta = datetime.combine(filtro_fecha_hasta, datetime.max.time(), tzinfo=timezone.get_current_timezone())
            locales_negocio_qs = locales_negocio_qs.filter(fecha_respuesta__lte=dt_hasta)
        if filtro_fecha_exacta:
            dt_exacta = datetime.combine(filtro_fecha_exacta, datetime.min.time(), tzinfo=timezone.get_current_timezone())
            dt_exacta_fin = datetime.combine(filtro_fecha_exacta, datetime.max.time(), tzinfo=timezone.get_current_timezone())
            locales_negocio_qs = locales_negocio_qs.filter(fecha_respuesta__range=(dt_exacta, dt_exacta_fin))

        locales_negocio_total = locales_negocio_qs.count()
        ctx['kpi_total_respuestas'] = locales_negocio_total

        # Aplicar rango temporal (si no hay filtros personalizados de fecha, usamos temp)
        if not (filtro_fecha_desde or filtro_fecha_hasta or filtro_fecha_exacta):
            locales_rango = locales_negocio_qs.filter(fecha_respuesta__gte=desde_temp)
        else:
            locales_rango = locales_negocio_qs
        ctx['kpi_respuestas_rango'] = locales_rango.count()
        ctx['kpi_respuestas_30d'] = locales_rango.count()

        # Filtrar por categoría NPS
        if filtro_nps_cat == 'PROMOTOR':
            locales_rango = locales_rango.filter(nps_puntaje__gte=9)
        elif filtro_nps_cat == 'PASIVO':
            locales_rango = locales_rango.filter(nps_puntaje__in=[7, 8])
        elif filtro_nps_cat == 'DETRACTOR':
            locales_rango = locales_rango.filter(nps_puntaje__lte=6)

        if locales_negocio_total:
            promedio_nps = locales_negocio_qs.exclude(nps_puntaje__isnull=True).aggregate(avg=Avg('nps_puntaje'))['avg'] or 0
            promotores = locales_rango.filter(nps_puntaje__gte=9).count()
            pasivos = locales_rango.filter(nps_puntaje__in=[7, 8]).count()
            detractores = locales_rango.filter(nps_puntaje__lte=6).count()
            total_cat = promotores + pasivos + detractores or 1
            nps_score = round(100 * (promotores - detractores) / total_cat)
        else:
            promedio_nps = 0
            promotores = pasivos = detractores = 0
            total_cat = 1
            nps_score = 0
        ctx['kpi_nps_promedio'] = round(promedio_nps, 1) if promedio_nps else 0
        ctx['kpi_nps_score'] = nps_score
        ctx['kpi_promotores'] = promotores
        ctx['kpi_pasivos'] = pasivos
        ctx['kpi_detractores'] = detractores
        if nps_score >= 50:
            ctx['nps_color'] = 'emerald'
            ctx['nps_badge'] = 'Promotor'
        elif nps_score >= 0:
            ctx['nps_color'] = 'amber'
            ctx['nps_badge'] = 'Pasivo'
        else:
            ctx['nps_color'] = 'rose'
            ctx['nps_badge'] = 'Detractor'

        if locales_negocio_total:
            muy_feliz = locales_rango.filter(csat_emocion=EmocionCSATChoices.MUY_FELIZ).count()
            feliz = locales_rango.filter(csat_emocion=EmocionCSATChoices.FELIZ).count()
            csat_total = locales_rango.exclude(csat_emocion__isnull=True).count() or 1
            ctx['kpi_csat_felices_pct'] = round(100 * (muy_feliz + feliz) / csat_total) if csat_total else 0
            ctx['kpi_csat_total'] = csat_total
            ctx['kpi_muy_feliz'] = muy_feliz
            ctx['kpi_feliz'] = feliz
            ctx['kpi_neutral'] = locales_rango.filter(csat_emocion=EmocionCSATChoices.NEUTRAL).count()
            ctx['kpi_insatisfecho'] = locales_rango.filter(csat_emocion=EmocionCSATChoices.INSATISFECHO).count()
            ctx['kpi_muy_insatisfecho'] = locales_rango.filter(csat_emocion=EmocionCSATChoices.MUY_INSATISFECHO).count()
        else:
            ctx['kpi_csat_felices_pct'] = 0
            ctx['kpi_csat_total'] = 0
            ctx['kpi_muy_feliz'] = 0
            ctx['kpi_feliz'] = 0
            ctx['kpi_neutral'] = 0
            ctx['kpi_insatisfecho'] = 0
            ctx['kpi_muy_insatisfecho'] = 0

        ultimas_respuestas = list(locales_rango.order_by('-fecha_respuesta')[:20])
        ctx['ultimas_respuestas'] = ultimas_respuestas

        locales_primer_local = negocio.locales.order_by('fecha_creacion').first()
        if locales_primer_local:
            ctx['qr_primer_local'] = locales_primer_local
            ctx['qr_url'] = request.build_absolute_uri('/e/' + locales_primer_local.qr_token + '/')

        # ==========================================================
        #  ANÁLISIS LATENCIA + DISTRIBUCIÓN TEMPORAL
        # ==========================================================
        analisis_latencia = None
        if locales_negocio_total >= 2:
            try:
                fechas = list(
                    locales_rango.order_by('fecha_respuesta')
                    .values_list('fecha_respuesta', flat=True)
                )
                if len(fechas) >= 2:
                    deltas_horas = []
                    for i in range(1, len(fechas)):
                        d = (fechas[i] - fechas[i-1]).total_seconds() / 3600.0
                        if d >= 0:
                            deltas_horas.append(d)
                    if deltas_horas:
                        prom_horas = sum(deltas_horas) / len(deltas_horas)
                        ctx['analisis_latencia'] = {
                            'n_intervalos': len(deltas_horas),
                            'promedio_horas': round(prom_horas, 1),
                            'promedio_minutos': round(prom_horas * 60, 0),
                            'respuestas_por_dia_prom': round(len(fechas) / max(1, temp_days), 2),
                            'primera': fechas[0],
                            'ultima': fechas[-1],
                        }
                # Distribución hora del día
                horas_count = [0] * 24
                dias_semana_count = [0] * 7
                dias_labels = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
                for dt in fechas:
                    h = timezone.localtime(dt).hour
                    horas_count[h] += 1
                    d = dt.weekday()  # 0=lun a 6=dom
                    dias_semana_count[d] += 1
                ctx['dist_horas'] = {
                    'labels': list(range(24)),
                    'data': horas_count,
                    'pico_hora': max(range(24), key=lambda i: horas_count[i]),
                }
                ctx['dist_dias'] = {
                    'labels': dias_labels,
                    'data': dias_semana_count,
                    'pico_dia': dias_labels[max(range(7), key=lambda i: dias_semana_count[i])] if any(dias_semana_count) else '—',
                }
            except Exception as e:
                ctx['analisis_latencia_error'] = str(e)

        # ==========================================================
        #  ANÁLISIS TEMÁTICA 4D (palabras clave)
        # ==========================================================
        textos_comentarios = []
        for r in locales_rango:
            c = getattr(r, 'comentario', None) or ''
            if c:
                textos_comentarios.append(c)
        tematicas, total_menciones = _analizar_tematicas(textos_comentarios)
        ctx['tematicas_4d'] = tematicas
        ctx['tematicas_total_menciones'] = total_menciones
        ctx['comentarios_analizados'] = len(textos_comentarios)

        # ==========================================================
        #  RESEÑAS GOOGLE (estrellas 1-5 con diferenciador)
        # ==========================================================
        from apps.reputation.models import ResenaGoogle, SentimientoChoices
        google_qs = ResenaGoogle.objects.filter(
            local__negocio=negocio
        ).select_related('local').order_by('-fecha_google')

        if filtro_local_id:
            try:
                lid = int(filtro_local_id)
                google_qs = google_qs.filter(local_id=lid)
            except (ValueError, TypeError):
                pass

        if not (filtro_fecha_desde or filtro_fecha_hasta or filtro_fecha_exacta):
            google_rango = google_qs.filter(fecha_google__gte=desde_temp)
        else:
            google_rango = google_qs
        if filtro_fecha_desde:
            dt_desde = datetime.combine(filtro_fecha_desde, datetime.min.time(), tzinfo=timezone.get_current_timezone())
            google_rango = google_rango.filter(fecha_google__gte=dt_desde)
        if filtro_fecha_hasta:
            dt_hasta = datetime.combine(filtro_fecha_hasta, datetime.max.time(), tzinfo=timezone.get_current_timezone())
            google_rango = google_rango.filter(fecha_google__lte=dt_hasta)
        if filtro_fecha_exacta:
            dt1 = datetime.combine(filtro_fecha_exacta, datetime.min.time(), tzinfo=timezone.get_current_timezone())
            dt2 = datetime.combine(filtro_fecha_exacta, datetime.max.time(), tzinfo=timezone.get_current_timezone())
            google_rango = google_rango.filter(fecha_google__range=(dt1, dt2))

        ctx['google_total'] = google_qs.count()
        ctx['google_rango'] = google_rango.count()
        google_rango_list = list(google_rango[:30])
        if google_qs.exists():
            rating_avg = google_qs.aggregate(avg=Avg('calificacion'))['avg'] or 0
            ctx['google_rating_promedio'] = round(float(rating_avg), 1)
            ctx['google_estrellas'] = _estrellas_html(rating_avg)
            rating_counts = {}
            for s in range(1, 6):
                rating_counts[s] = google_qs.filter(calificacion=s).count()
            ctx['google_rating_dist'] = rating_counts
        else:
            ctx['google_rating_promedio'] = 0
            ctx['google_estrellas'] = _estrellas_html(0)
            ctx['google_rating_dist'] = {1:0, 2:0, 3:0, 4:0, 5:0}
        ctx['google_resenas'] = google_rango_list

        rating_dist = ctx.get('google_rating_dist', {1:0,2:0,3:0,4:0,5:0})
        google_total_count = ctx.get('google_total', 0) or 1
        google_rating_rows = []
        for stars in (5, 4, 3, 2, 1):
            c = rating_dist.get(stars, 0)
            google_rating_rows.append({
                'stars': stars,
                'estrellas': _estrellas_html(stars),
                'count': c,
                'pct': round(100 * c / google_total_count, 0) if google_total_count else 0,
            })
        ctx['google_rating_rows'] = google_rating_rows

        if 'dist_horas' in ctx:
            horas_data = ctx['dist_horas']['data']
            ctx['dist_horas']['max_val'] = max(horas_data) if any(horas_data) else 1
        if 'dist_dias' in ctx:
            dias_max = max(ctx['dist_dias']['data']) if any(ctx['dist_dias']['data']) else 1
            ctx['dist_dias']['max_val'] = dias_max
            dias_list = []
            for i, lbl in enumerate(ctx['dist_dias']['labels']):
                val = ctx['dist_dias']['data'][i]
                dias_list.append({
                    'label': lbl,
                    'count': val,
                    'pct': round(100 * val / dias_max, 0) if dias_max else 0,
                })
            ctx['dist_dias']['rows'] = dias_list

        tematicas_list = list(tematicas.items())
        tematicas_list.sort(key=lambda kv: kv[1]['menciones'], reverse=True)
        ctx['tematicas_ranking'] = tematicas_list

        # ==========================================================
        #  BENCHMARK (datos por Local + Top competidores)
        # ==========================================================
        from apps.reputation.models import Benchmark
        benchmark_locales_data = []
        locales_negocio = negocio.locales.filter(estado='ACTIVO').order_by('fecha_creacion')
        total_locales = locales_negocio.count()
        locales_con_benchmark = 0
        rating_promedio_acum = 0.0
        posicion_promedio_acum = 0
        diferencia_vs_rubro_acum = 0.0
        benchmark_local_activo_id = None
        try:
            benchmark_local_activo_id = int(filtro_local_id) if filtro_local_id else None
        except (ValueError, TypeError):
            benchmark_local_activo_id = None

        for local_obj in locales_negocio:
            ult_b = Benchmark.objects.filter(local=local_obj).order_by('-fecha_generacion').first()
            entry = {
                'local_id': local_obj.id,
                'local_nombre': local_obj.nombre,
                'tiene_benchmark': ult_b is not None,
                'benchmark': ult_b,
                'fecha_generacion': ult_b.fecha_generacion if ult_b else None,
                'rating_local': float(ult_b.puntuacion_local) if ult_b else None,
                'posicion_local': ult_b.posicion_local if ult_b else None,
                'total_evaluados': ult_b.total_evaluados if ult_b else None,
                'promedio_rubro': float(ult_b.puntuacion_promedio_rubro) if ult_b else None,
                'top25_promedio': float(ult_b.top25_promedio) if ult_b else None,
                'bottom25_promedio': float(ult_b.bottom25_promedio) if ult_b else None,
                'percentiles': {},
                'competidores_rows': [],
                'diferencia_vs_rubro': None,
                'delta_label': None,
                'delta_color': 'slate',
            }
            if ult_b:
                locales_con_benchmark += 1
                rating_promedio_acum += entry['rating_local']
                posicion_promedio_acum += entry['posicion_local']
                diff = entry['rating_local'] - entry['promedio_rubro']
                entry['diferencia_vs_rubro'] = round(diff, 2)
                if diff > 0:
                    entry['delta_label'] = "+%.2f vs rubro" % entry['diferencia_vs_rubro']
                    entry['delta_color'] = 'emerald'
                elif diff < 0:
                    entry['delta_label'] = "%.2f vs rubro" % entry['diferencia_vs_rubro']
                    entry['delta_color'] = 'rose'
                else:
                    entry['delta_label'] = '= al promedio rubro'
                    entry['delta_color'] = 'amber'
                diferencia_vs_rubro_acum += diff
                datos_b = ult_b.datos or {}
                entry['percentiles'] = datos_b.get('percentiles', {}) or {}
                competidores_raw = datos_b.get('competidores', []) or []
                comp_rows = []
                for idx, c in enumerate(competidores_raw[:10], start=1):
                    puntuacion = c.get('puntuacion') or c.get('rating') or 0
                    opiniones = c.get('opiniones') or c.get('numero_opiniones') or 0
                    try:
                        puntuacion_f = round(float(puntuacion), 1)
                    except Exception:
                        puntuacion_f = 0.0
                    try:
                        opiniones_i = int(opiniones)
                    except Exception:
                        opiniones_i = 0
                    pct_bar = int(max(0, min(100, (puntuacion_f / 5.0) * 100))) if puntuacion_f else 0
                    if entry['rating_local'] and puntuacion_f >= entry['rating_local']:
                        badge = 'Top competidor'
                        badge_color = 'amber'
                    elif entry['promedio_rubro'] and puntuacion_f >= entry['promedio_rubro']:
                        badge = 'Sobre promedio'
                        badge_color = 'blue'
                    else:
                        badge = 'Bajo promedio'
                        badge_color = 'slate'
                    comp_rows.append({
                        'pos': idx,
                        'nombre': c.get('nombre') or c.get('place_nombre') or 'Competidor',
                        'puntuacion': puntuacion_f,
                        'opiniones': opiniones_i,
                        'pct_bar': pct_bar,
                        'badge': badge,
                        'badge_color': badge_color,
                        'direccion': c.get('direccion') or c.get('place_direccion') or '',
                        'place_id': c.get('place_id') or '',
                    })
                entry['competidores_rows'] = comp_rows
                if benchmark_local_activo_id is None:
                    benchmark_local_activo_id = local_obj.id
            benchmark_locales_data.append(entry)

        ctx['benchmark_locales'] = benchmark_locales_data
        ctx['benchmark_total_locales'] = total_locales
        ctx['benchmark_locales_con_benchmark'] = locales_con_benchmark

        if locales_con_benchmark > 0:
            ctx['benchmark_rating_promedio'] = round(rating_promedio_acum / locales_con_benchmark, 1)
            ctx['benchmark_posicion_promedio'] = round(posicion_promedio_acum / locales_con_benchmark, 1)
            ctx['benchmark_diferencia_vs_rubro_promedio'] = round(diferencia_vs_rubro_acum / locales_con_benchmark, 2)
        else:
            ctx['benchmark_rating_promedio'] = 0
            ctx['benchmark_posicion_promedio'] = 0
            ctx['benchmark_diferencia_vs_rubro_promedio'] = 0

        ctx['benchmark_local_activo_id'] = benchmark_local_activo_id
        activo_entry = None
        for e in benchmark_locales_data:
            if e['local_id'] == benchmark_local_activo_id:
                activo_entry = e
                break
        if activo_entry is None and benchmark_locales_data:
            activo_entry = benchmark_locales_data[0]
        ctx['benchmark_activo'] = activo_entry

        # ==========================================================
        #  EXPORT EXCEL
        # ==========================================================
        puede_exportar = request.user.is_superuser or (
            ctx.get('plan_activo') and getattr(ctx['plan_activo'], 'tiene_export_excel', False)
        )
        ctx['puede_exportar_excel'] = puede_exportar

        if request.GET.get('export') == 'excel' and puede_exportar:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse

            hoy = timezone.localdate()
            filename = f"ClientBeat-KPIs-{negocio.slug}-{hoy:%Y%m%d}.xlsx"

            wb = Workbook()
            delgada = Side(style='thin', color='D1D5DB')
            borde_completo = Border(left=delgada, right=delgada, top=delgada, bottom=delgada)
            fuente_header = Font(bold=True, color='FFFFFF', size=11)
            fill_header = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            fill_subhead = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
            centro = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws_kpi = wb.active
            ws_kpi.title = "Resumen KPI"
            ws_kpi['A1'] = f"ClientBeat - Reporte KPI {negocio.nombre}"
            ws_kpi['A1'].font = Font(bold=True, size=16, color='4F46E5')
            ws_kpi.merge_cells('A1:F1')
            ws_kpi['A2'] = f"Periodo: {temp_label} · Ubicación: {ubi_label} · Generado: {hoy:%d-%m-%Y}"
            ws_kpi['A2'].font = Font(size=11, italic=True, color='6B7280')
            ws_kpi.merge_cells('A2:F2')

            headers1 = ['Indicador', 'Valor', 'Detalle', 'Fórmula / Contexto', '', '']
            for col_idx, h in enumerate(headers1, start=1):
                celda = ws_kpi.cell(row=4, column=col_idx, value=h)
                celda.font = fuente_header
                celda.fill = fill_header
                celda.alignment = centro
                celda.border = borde_completo

            kpi_rows = [
                ("TOTAL RESPUESTAS (histórico)", locales_negocio_total, "", "Todas las RespuestaEncuesta del negocio", "", ""),
                (f"RESPUESTAS ÚLTIMOS {temp_days} DÍAS", ctx['kpi_respuestas_rango'], f"Rango filtrado {temp_label}", "", "", ""),
                ("", "", "", "", "", ""),
                ("== NPS (Net Promoter Score) ==", "", "", "Escala 0-10 · Promotores 9-10 · Pasivos 7-8 · Detractores 0-6", "", ""),
                ("NPS SCORE (RANGO)", nps_score, f"[{ctx['nps_badge']}]", "(%PROMOTORES - %DETRACTORES) · Bueno >= +50", "", ""),
                ("NPS Promedio Histórico (0-10)", ctx['kpi_nps_promedio'], "", "Media aritmética nps_puntaje (todas)", "", ""),
                ("  · Promotores (P)", promotores, "", "nps_puntaje 9 ó 10", "", ""),
                ("  · Pasivos (Pa)", pasivos, "", "nps_puntaje 7 ó 8", "", ""),
                ("  · Detractores (D)", detractores, "", "nps_puntaje 0 a 6", "", ""),
                ("", "", "", "", "", ""),
                ("== CSAT (Customer Satisfaction) ==", "", "", "Índice = Σ Felices / Total * 100", "", ""),
                ("CSAT Índice Felices (%)", ctx['kpi_csat_felices_pct'], "", "(Muy Feliz + Feliz) / CSAT Total * 100", "", ""),
                ("CSAT Total (con emoción)", ctx['kpi_csat_total'], "", "Respuestas con csat_emocion NO null", "", ""),
                ("  · Muy Feliz", ctx['kpi_muy_feliz'], "", "EmocionCSATChoices.MUY_FELIZ", "", ""),
                ("  · Feliz", ctx['kpi_feliz'], "", "EmocionCSATChoices.FELIZ", "", ""),
                ("  · Neutral", ctx['kpi_neutral'], "", "EmocionCSATChoices.NEUTRAL", "", ""),
                ("  · Insatisfecho", ctx['kpi_insatisfecho'], "", "EmocionCSATChoices.INSATISFECHO", "", ""),
                ("  · Muy Insatisfecho", ctx['kpi_muy_insatisfecho'], "", "EmocionCSATChoices.MUY_INSATISFECHO", "", ""),
                ("", "", "", "", "", ""),
                ("== RESEÑAS GOOGLE ==", "", "", "", "", ""),
                ("Total reseñas Google", ctx.get('google_total', 0), "", "Todas", "", ""),
                ("Rating promedio Google", ctx.get('google_rating_promedio', 0), "/ 5.0", "Calificación promedio", "", ""),
                ("", "", "", "", "", ""),
                ("== PLAN ACTIVO ==", "", "", "", "", ""),
                ("Plan", str(ctx.get('plan_activo') or "MVP Básico"), "", "Suscripcion ACTIVA del negocio", "", ""),
            ]
            for r_idx, fila in enumerate(kpi_rows, start=5):
                for c_idx, val in enumerate(fila, start=1):
                    cel = ws_kpi.cell(row=r_idx, column=c_idx, value=val)
                    cel.border = borde_completo
                    if isinstance(val, str) and val.startswith("=="):
                        cel.fill = fill_subhead
                        cel.font = Font(bold=True, color='4F46E5')
            anchos1 = [44, 14, 28, 50, 8, 8]
            for i, w in enumerate(anchos1, start=1):
                ws_kpi.column_dimensions[get_column_letter(i)].width = w

            ws_hist = wb.create_sheet(title="Histórico Respuestas")
            headers2 = [
                "#", "Fecha Respuesta", "Local", "RUT Cliente",
                "NPS 0-10", "Categoría NPS", "CSAT Emoción",
                "Edad", "Género", "¿Volverías?", "Comentario",
            ]
            for col_idx, h in enumerate(headers2, start=1):
                celda = ws_hist.cell(row=1, column=col_idx, value=h)
                celda.font = fuente_header
                celda.fill = fill_header
                celda.alignment = centro
                celda.border = borde_completo

            def nps_cat(p):
                if p is None: return ""
                if p >= 9: return "PROMOTOR"
                if p >= 7: return "PASIVO"
                return "DETRACTOR"

            for r_idx, resp in enumerate(locales_rango.order_by('fecha_respuesta'), start=2):
                fila = [
                    r_idx - 1,
                    timezone.localtime(resp.fecha_respuesta).strftime("%Y-%m-%d %H:%M") if resp.fecha_respuesta else "",
                    getattr(resp.local, 'nombre', '') if resp.local else "",
                    getattr(resp, 'cliente_rut', '') or "",
                    resp.nps_puntaje if resp.nps_puntaje is not None else "",
                    nps_cat(resp.nps_puntaje),
                    resp.get_csat_emocion_display() if resp.csat_emocion else "",
                    getattr(resp, 'get_rango_edad_display', lambda: '')() or "",
                    getattr(resp, 'get_genero_display', lambda: '')() or "",
                    getattr(resp, 'get_volveria_display', lambda: '')() or "",
                    (getattr(resp, 'comentario', None) or "")[:300],
                ]
                for c_idx, val in enumerate(fila, start=1):
                    celda = ws_hist.cell(row=r_idx, column=c_idx, value=val)
                    celda.border = borde_completo
                    celda.alignment = Alignment(vertical='center', wrap_text=True)
                    if c_idx == 6 and isinstance(val, str):
                        if val == "PROMOTOR": celda.font = Font(bold=True, color='059669')
                        elif val == "DETRACTOR": celda.font = Font(bold=True, color='DC2626')
                        elif val == "PASIVO": celda.font = Font(bold=True, color='D97706')
            anchos2 = [5, 19, 26, 14, 9, 12, 18, 14, 12, 14, 42]
            for i, w in enumerate(anchos2, start=1):
                ws_hist.column_dimensions[get_column_letter(i)].width = w
            ws_hist.freeze_panes = 'A2'

            ws_google = wb.create_sheet(title="Reseñas Google")
            headers_g = ["#", "Fecha Google", "Local", "Autor", "Calificación (★)", "Sentimiento", "Comentario"]
            for col_idx, h in enumerate(headers_g, start=1):
                celda = ws_google.cell(row=1, column=col_idx, value=h)
                celda.font = fuente_header
                celda.fill = fill_header
                celda.alignment = centro
                celda.border = borde_completo
            for r_idx, rg in enumerate(google_rango_list, start=2):
                fila_g = [
                    r_idx - 1,
                    timezone.localtime(rg.fecha_google).strftime("%Y-%m-%d") if rg.fecha_google else "",
                    getattr(rg.local, 'nombre', '') if rg.local else "",
                    rg.autor_nombre,
                    int(rg.calificacion or 0),
                    rg.get_sentimiento_display() if rg.sentimiento else "—",
                    (rg.comentario or "")[:500],
                ]
                for c_idx, val in enumerate(fila_g, start=1):
                    celda = ws_google.cell(row=r_idx, column=c_idx, value=val)
                    celda.border = borde_completo
                    celda.alignment = Alignment(vertical='center', wrap_text=True)
            anchos_g = [5, 14, 24, 22, 16, 14, 70]
            for i, w in enumerate(anchos_g, start=1):
                ws_google.column_dimensions[get_column_letter(i)].width = w
            ws_google.freeze_panes = 'A2'

            virtual_book = BytesIO()
            wb.save(virtual_book)
            virtual_book.seek(0)

            response = HttpResponse(
                virtual_book.read(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

    return render(request, 'dashboard_placeholder.html', ctx)


def legal(request, slug):
    data = _CONTENIDOS_LEGALES.get(slug)
    if not data:
        data = _CONTENIDOS_LEGALES['terminos']
    ctx = {
        'titulo': data['titulo'],
        'fecha_actualizacion': data['fecha'],
        'contenido': data['contenido'],
    }
    return render(request, 'legal.html', ctx)
